# -*- coding: utf-8 -*-
"""
DOKÜMAN OLUŞTURMA MODÜLÜ
========================
"""

import os
from datetime import datetime
from typing import List, Optional
from pathlib import Path

# PDF oluşturma
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm, mm
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_JUSTIFY
from reportlab.lib.colors import HexColor, black, grey, lightgrey, white
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    PageBreak, Image, HRFlowable
)
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Excel oluşturma
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# PDF birleştirme
from pypdf import PdfReader, PdfWriter

from config import (
    SourceArticle, CitingArticle, CandidateInfo,
    OUTPUT_DIR, sanitize_filename
)


class StyleManager:
    """PDF stilleri yöneticisi"""
    
    @staticmethod
    def register_fonts():
        # DejaVu fonts
        project_fonts = Path(__file__).parent / "fonts" / "dejavu-fonts-ttf-2.37" / "ttf"
        
        try:
            pdfmetrics.registerFont(TTFont('DejaVuSerif', str(project_fonts / "DejaVuSerif.ttf")))
            pdfmetrics.registerFont(TTFont('DejaVuSerif-Bold', str(project_fonts / "DejaVuSerif-Bold.ttf")))
            return True
        except:
            return False

    @staticmethod
    def get_styles() -> dict:
        StyleManager.register_fonts()
        font_name = 'DejaVuSerif'
        
        return {
            'normal': ParagraphStyle(
                'normal',
                fontName=font_name,
                fontSize=12,
                leading=16,
                spaceAfter=6,
                textColor=black
            ),
            'heading1': ParagraphStyle(
                'heading1',
                fontName=font_name + '-Bold', # Use bold variant
                fontSize=12,
                leading=16,
                spaceAfter=12,
                textColor=black
            ),
             'title': ParagraphStyle(
                'title',
                fontName=font_name + '-Bold',
                fontSize=14,
                leading=18,
                spaceAfter=20,
                alignment=TA_CENTER,
                textColor=black
            ),
        }


class CitationDocumentBuilder:
    """Atıf dokümanı oluşturucu"""
    
    def __init__(self, candidate: CandidateInfo, source_article: SourceArticle):
        self.candidate = candidate
        self.source = source_article
        self.styles = StyleManager.get_styles()
        self.page_number = 0
    
    def _add_page_number(self, canvas, doc):
        """Sayfa numarası ekle"""
        self.page_number += 1
        canvas.saveState()
        canvas.setFont('DejaVuSerif', 10)
        canvas.setFillColor(black)
        # Sağ alt köşe
        canvas.drawRightString(A4[0] - 2*cm, 1.5*cm, str(self.page_number))
        canvas.restoreState()
    
    def _create_header_section(self) -> List:
        """Eser başlığı ve bilgileri"""
        elements = []
        
        # Eser Adı
        # Format: Eser Adı: [Title] (Eser Yök İd:[ID])
        title_text = f"<b>Eser Adı:</b> {self.source.title}"
        if self.source.wos_id:
             title_text += f" (Eser Yök İd:{self.source.wos_id})"
        
        elements.append(Paragraph(title_text, self.styles['normal']))
        elements.append(Spacer(1, 0.5*cm))
        
        # Atıflar Başlığı
        elements.append(Paragraph("<b>Atıflar</b>", self.styles['normal']))
        
        # Atıf Listesi
        for i, article in enumerate(self.source.citing_articles, 1):
            # Format: 1. Author, et al. "Title" Journal Vol.Issue (Year): Pages.
            citation_text = f"{i}. {article.get_citation_string('apa')}"
            elements.append(Paragraph(citation_text, self.styles['normal']))
            
        elements.append(PageBreak())
        return elements

    def _create_citation_separator_page(self, index: int, article: CitingArticle) -> List:
        """Atıf ayracı sayfası (Atıf 1, 2, 3...)"""
        elements = []
        
        # Başlık: Atıf N
        elements.append(Paragraph(f"<b>Atıf {index}</b>", self.styles['title']))
        elements.append(Spacer(1, 1.0*cm))
        
        # Makale bilgisi
        citation_info = article.get_citation_string()
        elements.append(Paragraph(citation_info, self.styles['normal']))
        
        elements.append(PageBreak())
        return elements

    def build_pdf(self, output_path: str) -> bool:
        """
        Base PDF dokümanını oluştur (Listeler ve Ayraçlar)
        """
        try:
            doc = SimpleDocTemplate(
                output_path,
                pagesize=A4,
                rightMargin=2.5*cm,
                leftMargin=2.5*cm,
                topMargin=2.5*cm,
                bottomMargin=2.5*cm
            )
            
            elements = []
            
            # 1. Giriş Sayfası (Eser Adı ve Atıf Listesi)
            elements.extend(self._create_header_section())
            
            # 2. Her atıf için ayraç sayfası
            # 2. Her atıf için ayraç sayfası
            for i, article in enumerate(self.source.citing_articles, 1):
                elements.extend(self._create_citation_separator_page(i, article))
            
            # PDF oluştur
            doc.build(
                elements,
                onFirstPage=self._add_page_number,
                onLaterPages=self._add_page_number
            )
            
            print(f"✅ Base PDF oluşturuldu: {output_path}")
            return True
            
        except Exception as e:
            print(f"❌ PDF oluşturma hatası: {e}")
            import traceback
            traceback.print_exc()
            return False

    def build_excel(self, output_path: str) -> bool:
        try:
            wb = Workbook()
            ws_citations = wb.active
            ws_citations.title = "Atıf Listesi"
            
            headers = ["#", "Başlık", "Yazarlar", "Dergi", "Yıl", "DOI"]
            ws_citations.append(headers)
            
            for i, article in enumerate(self.source.citing_articles, 1):
                row = [
                    i,
                    article.title,
                    ", ".join(article.authors),
                    article.journal,
                    article.year,
                    article.doi
                ]
                ws_citations.append(row)
            
            wb.save(output_path)
            return True
        except:
            return False


class FinalDocumentAssembler:
    """
    Final doküman birleştirici
    """
    
    def __init__(self, source_article: SourceArticle):
        self.source = source_article
    
    def _highlight_page(self, page, bbox: List[float]):
        from io import BytesIO
        from reportlab.pdfgen import canvas as rl_canvas
        from reportlab.lib.colors import yellow
        from pypdf import PdfReader as PyPdfReader
        
        if not bbox: return page

        media_box = page.mediabox
        page_width = float(media_box.width)
        page_height = float(media_box.height)
        
        packet = BytesIO()
        c = rl_canvas.Canvas(packet, pagesize=(page_width, page_height))
        
        x0, top, x1, bottom = bbox
        y0 = page_height - bottom
        y1 = page_height - top
        
        c.setFillColor(yellow)
        c.setFillAlpha(0.3)
        c.rect(x0, y0, x1 - x0, y1 - y0, fill=True, stroke=False)
        c.save()
        
        packet.seek(0)
        overlay_reader = PyPdfReader(packet)
        overlay_page = overlay_reader.pages[0]
        
        page.merge_page(overlay_page)
        return page
    
    def _create_section_header(self, writer, text: str):
        """Bölüm başlığı sayfası oluştur"""
        from io import BytesIO
        from reportlab.pdfgen import canvas as rl_canvas
        from reportlab.lib.pagesizes import A4
        
        packet = BytesIO()
        c = rl_canvas.Canvas(packet, pagesize=A4)
        
        # DejaVu font kullan
        try:
            c.setFont('DejaVuSerif-Bold', 14)
        except:
            c.setFont('Helvetica-Bold', 14)
        
        # Başlığı sayfanın üst kısmına yaz
        c.drawString(2.5*cm, A4[1] - 2.5*cm, text)
        c.save()
        
        packet.seek(0)
        from pypdf import PdfReader as PyPdfReader
        header_reader = PyPdfReader(packet)
        writer.add_page(header_reader.pages[0])
    
    def _add_image_as_page(self, writer, image_path: str):
        """Convert an image file to a PDF page and add to writer"""
        from io import BytesIO
        from reportlab.pdfgen import canvas as rl_canvas
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.utils import ImageReader
        from pypdf import PdfReader as PyPdfReader
        
        try:
            # Create PDF page from image
            packet = BytesIO()
            c = rl_canvas.Canvas(packet, pagesize=A4)
            
            # Get image dimensions
            img = ImageReader(image_path)
            img_width, img_height = img.getSize()
            
            # Scale to fit A4 page with margins
            page_width, page_height = A4
            margin = 50
            max_width = page_width - 2 * margin
            max_height = page_height - 2 * margin
            
            # Calculate scale factor
            scale_w = max_width / img_width
            scale_h = max_height / img_height
            scale = min(scale_w, scale_h, 1.0)  # Don't scale up
            
            new_width = img_width * scale
            new_height = img_height * scale
            
            # Center on page
            x = (page_width - new_width) / 2
            y = (page_height - new_height) / 2
            
            c.drawImage(img, x, y, width=new_width, height=new_height)
            c.save()
            
            packet.seek(0)
            img_reader = PyPdfReader(packet)
            writer.add_page(img_reader.pages[0])
        except Exception as e:
            print(f"Error converting image to PDF: {e}")
    
    def assemble(self, base_pdf_path: str, output_path: str) -> bool:
        """
        Base PDF (Liste + Ayraçlar) ile Tam Metin Sayfalarını Birleştir
        
        Her atıf için yapı:
        1. Atıf N (separator page)
        2. A.N. Yayının Ünvan Sayfası (kullanıcı kapak görselleri)
        3. AN. Eserin Başlık Sayfası (PDF sayfa 1)
        4. AN. Eserde atıf yapılan sayfalar (TÜM atıf sayfaları, sarı highlight)
        5. AN. Kaynakça Sayfası (referans sayfası, sarı highlight)
        """
        try:
            writer = PdfWriter()
            
            if not os.path.exists(base_pdf_path):
                return False
                
            base_reader = PdfReader(base_pdf_path)
            base_pages = list(base_reader.pages)
            
            num_citations = len(self.source.citing_articles)
            total_base_pages = len(base_pages)
            num_list_pages = total_base_pages - num_citations
            
            if num_list_pages < 1: num_list_pages = 1
            
            # 1. Add List Pages (Atıf Listesi)
            for i in range(num_list_pages):
                if i < len(base_pages):
                    writer.add_page(base_pages[i])
            
            separator_start_index = num_list_pages
            
            # 2. Her atıf için içerik ekle
            for i, article in enumerate(self.source.citing_articles):
                index = i + 1
                
                # A. Separator Page (Atıf N)
                sep_idx = separator_start_index + i
                if sep_idx < len(base_pages):
                    writer.add_page(base_pages[sep_idx])
                
                # PDF yoksa devam et
                if not article.pdf_path or not os.path.exists(article.pdf_path):
                    continue
                
                try:
                    reader = PdfReader(article.pdf_path)
                    
                    # B. A.N. Yayının Ünvan Sayfası (Kapak Görselleri)
                    self._create_section_header(writer, f"A.{index}. Yayının Ünvan Sayfası (kitap, dergi, vb.)")
                    
                    # Kapak sayfalarını ekle - önce cover_pages_all listesini kontrol et
                    cover_added = False
                    
                    # Tüm kapak sayfalarını ekle (yeni sistem)
                    if article.cover_pages_all and len(article.cover_pages_all) > 0:
                        for cover_path in article.cover_pages_all:
                            if cover_path and os.path.exists(cover_path):
                                try:
                                    # Check file extension
                                    ext = cover_path.lower().split('.')[-1]
                                    if ext in ['png', 'jpg', 'jpeg', 'gif', 'bmp']:
                                        # Convert image to PDF page
                                        self._add_image_as_page(writer, cover_path)
                                        cover_added = True
                                    elif ext == 'pdf':
                                        # PDF file - read directly
                                        cover_reader = PdfReader(cover_path)
                                        for p in cover_reader.pages:
                                            writer.add_page(p)
                                            cover_added = True
                                except Exception as e:
                                    print(f"Error adding cover {cover_path}: {e}")
                    
                    # Fallback: Eski cover_page_path kullan
                    if not cover_added and article.cover_page_path and os.path.exists(article.cover_page_path):
                        try:
                            ext = article.cover_page_path.lower().split('.')[-1]
                            if ext in ['png', 'jpg', 'jpeg', 'gif', 'bmp']:
                                self._add_image_as_page(writer, article.cover_page_path)
                                cover_added = True
                            elif ext == 'pdf':
                                cover_reader = PdfReader(article.cover_page_path)
                                for p in cover_reader.pages:
                                    writer.add_page(p)
                                    cover_added = True
                        except Exception as e:
                            print(f"Error adding fallback cover: {e}")
                    
                    # Kapak bulunamadıysa, downloads klasöründe ara
                    if not cover_added:
                        kapak_path = os.path.join(os.path.dirname(article.pdf_path), f"kapak_{index}.pdf")
                        if os.path.exists(kapak_path):
                            try:
                                cover_reader = PdfReader(kapak_path)
                                for p in cover_reader.pages:
                                    writer.add_page(p)
                                    cover_added = True
                            except: pass
                    
                    # NOT: Kapak bulunamadığında artık PDF'in ilk sayfasını EKLEMİYORUZ
                    # Çünkü bu sayfa zaten "Eserin Başlık Sayfası" bölümünde gösterilecek
                    # Kullanıcı kapak yüklemediyse bu bölüm boş kalır - bu beklenen davranış
                    
                    # C. AN. Eserin Başlık Sayfası (PDF ilk sayfa)
                    self._create_section_header(writer, f"A{index}. Eserin Başlık Sayfası")
                    
                    if article.title_page and 1 <= article.title_page <= len(reader.pages):
                        writer.add_page(reader.pages[article.title_page - 1])
                    elif len(reader.pages) > 0:
                        writer.add_page(reader.pages[0])
                    
                    # D. AN. Eserde atıf yapılan sayfalar (TÜM sayfalar, highlight ile)
                    self._create_section_header(writer, f"A{index}. Eserde atıf yapılan sayfa(lar)")
                    
                    # citation_pages listesindeki TÜM sayfaları ekle
                    if article.citation_pages and len(article.citation_pages) > 0:
                        added_pages = set()
                        for page_idx, page_num in enumerate(article.citation_pages):
                            if page_num and 1 <= page_num <= len(reader.pages):
                                if page_num not in added_pages:
                                    page = reader.pages[page_num - 1]
                                    # Bu sayfa için bbox varsa highlight ekle
                                    if article.citation_bboxes and page_idx < len(article.citation_bboxes):
                                        try:
                                            page = self._highlight_page(page, article.citation_bboxes[page_idx])
                                        except: pass
                                    writer.add_page(page)
                                    added_pages.add(page_num)
                    
                    # E. AN. Kaynakça Sayfası (highlight ile)
                    self._create_section_header(writer, f"A{index}. Kaynakça Sayfası")
                    
                    if article.reference_page and 1 <= article.reference_page <= len(reader.pages):
                        page = reader.pages[article.reference_page - 1]
                        if article.reference_bbox:
                            try:
                                page = self._highlight_page(page, article.reference_bbox)
                            except: pass
                        writer.add_page(page)
                    
                except Exception as e:
                    print(f"Error processing PDF {article.pdf_path}: {e}")
            
            with open(output_path, 'wb') as f:
                writer.write(f)
                
            return True
            
        except Exception as e:
            print(f"Assemble Error: {e}")
            import traceback
            traceback.print_exc()
            return False